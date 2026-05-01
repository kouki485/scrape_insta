from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from src.excel import LEADS_HEADERS, LEADS_SHEET, SEEN_HEADERS, SEEN_SHEET, ExcelStore
from src.filter import Lead


def _lead(username: str, followers: int = 5000) -> Lead:
    now = datetime(2026, 5, 1, 8, 0, tzinfo=timezone.utc)
    return Lead(
        fetched_at=now,
        username=username,
        full_name=username,
        follower_count=followers,
        following_count=500,
        bio="EN bio",
        bio_lang="non-ja",
        latest_post_url=f"https://instagram.com/p/{username}",
        latest_post_caption="caption",
        caption_lang="non-ja",
        post_timestamp=now,
        matched_hashtag="asakusa",
        profile_url=f"https://instagram.com/{username}",
    )


def test_creates_workbook_with_headers_when_missing(tmp_path: Path):
    path = tmp_path / "out.xlsx"
    store = ExcelStore(path)
    store.save()

    wb = load_workbook(path)
    assert wb[LEADS_SHEET][1][0].value == LEADS_HEADERS[0]
    assert wb[SEEN_SHEET][1][0].value == SEEN_HEADERS[0]


def test_append_leads_writes_rows(tmp_path: Path):
    path = tmp_path / "out.xlsx"
    store = ExcelStore(path)
    appended = store.append_leads([_lead("a"), _lead("b", followers=12000)])
    store.save()

    assert appended == 2
    wb = load_workbook(path)
    ws = wb[LEADS_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert {r[1] for r in rows} == {"a", "b"}


def test_update_seen_records_users(tmp_path: Path):
    path = tmp_path / "out.xlsx"
    store = ExcelStore(path)
    store.update_seen(["alice", "bob"])
    store.save()

    wb = load_workbook(path)
    ws = wb[SEEN_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert {r[0] for r in rows} == {"alice", "bob"}


def test_load_seen_returns_users_within_window(tmp_path: Path):
    path = tmp_path / "out.xlsx"
    store = ExcelStore(path)
    now = datetime(2026, 5, 1, 8, 0, tzinfo=timezone.utc)
    store.update_seen(["recent_user"], now=now)
    store.update_seen(["old_user"], now=now - timedelta(days=10))
    store.save()

    fresh = ExcelStore(path)
    seen = fresh.load_seen(window_days=7, now=now)
    assert "recent_user" in seen
    assert "old_user" not in seen


def test_persists_across_reopens(tmp_path: Path):
    path = tmp_path / "out.xlsx"
    a = ExcelStore(path)
    a.append_leads([_lead("first")])
    a.save()

    b = ExcelStore(path)
    b.append_leads([_lead("second")])
    b.save()

    wb = load_workbook(path)
    rows = list(wb[LEADS_SHEET].iter_rows(min_row=2, values_only=True))
    assert {r[1] for r in rows} == {"first", "second"}
