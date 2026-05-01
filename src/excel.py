from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from src.filter import Lead

LEADS_HEADERS = (
    "fetched_at",
    "username",
    "full_name",
    "follower_count",
    "following_count",
    "bio",
    "bio_lang",
    "latest_post_url",
    "latest_post_caption",
    "caption_lang",
    "post_timestamp",
    "matched_hashtag",
    "profile_url",
)

SEEN_HEADERS = ("username", "first_seen_at")

LEADS_SHEET = "leads"
SEEN_SHEET = "seen_users"


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default = wb.active
    default.title = LEADS_SHEET
    default.append(list(LEADS_HEADERS))
    seen = wb.create_sheet(SEEN_SHEET)
    seen.append(list(SEEN_HEADERS))
    return wb


def _ensure_sheets(wb: Workbook) -> None:
    if LEADS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(LEADS_SHEET)
        ws.append(list(LEADS_HEADERS))
    if SEEN_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SEEN_SHEET)
        ws.append(list(SEEN_HEADERS))


class ExcelStore:
    def __init__(self, path: Path | str):
        self._path = Path(path)
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._wb = _ensure_workbook(self._path)
        _ensure_sheets(self._wb)

    @property
    def path(self) -> Path:
        return self._path

    def load_seen(self, *, window_days: int, now: datetime | None = None) -> set[str]:
        now = now or datetime.now(timezone.utc)
        cutoff = now - timedelta(days=window_days)
        ws = self._wb[SEEN_SHEET]
        seen: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            username = str(row[0]).strip()
            ts_raw = row[1]
            if not username or ts_raw is None:
                continue
            first_seen = _coerce_datetime(ts_raw)
            if first_seen is None:
                continue
            if first_seen >= cutoff:
                seen.add(username)
        return seen

    def append_leads(self, leads: Iterable[Lead]) -> int:
        ws = self._wb[LEADS_SHEET]
        count = 0
        for lead in leads:
            ws.append(_lead_to_row(lead))
            count += 1
        return count

    def update_seen(self, usernames: Iterable[str], *, now: datetime | None = None) -> int:
        now = now or datetime.now(timezone.utc)
        ws = self._wb[SEEN_SHEET]
        count = 0
        for username in usernames:
            if not username:
                continue
            ws.append([username, now.isoformat()])
            count += 1
        return count

    def save(self) -> None:
        self._wb.save(self._path)


def _coerce_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    return None


def _lead_to_row(lead: Lead) -> list[object]:
    return [
        lead.fetched_at.isoformat(),
        lead.username,
        lead.full_name,
        lead.follower_count,
        lead.following_count,
        lead.bio,
        lead.bio_lang,
        lead.latest_post_url,
        lead.latest_post_caption,
        lead.caption_lang,
        lead.post_timestamp.isoformat(),
        lead.matched_hashtag,
        lead.profile_url,
    ]
